#first install from cmd-pip install openpyxl, this is pyxl package/library, this will be
#installed in the local interpreter level hence make sure interpreter is selected as the local
#one and not the pycharm one.

import openpyxl
#this is to store the excel location
book =openpyxl.load_workbook("https:\\people.ey.com\\personal\\jeet_chakraborty1_gds_ey_com\\Documents\\Desktop\\TestData.xlsx?web=1")
#this is to get the active sheet
sheet =book.active
Dict = {}
#this is to ge the cell
cell =sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Rahul"#to write data in excel

print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)#to get the number of rows

print(sheet.max_column)#to get the number of columns

print(sheet['A5'].value)

#logic to fetch data
for i in range(1,sheet.max_row+1):  # to get rows
    if sheet.cell(row =i,column=1).value == "Testcase2":

        for j in range(2,sheet.max_column+1):#to get columns
            #Dict["lastname"]="shetty
            Dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i, column=j).value

print(Dict)