import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Rahul","lastname":"shetty","gender":"Male"}, {"firstname":"Anshika", "lastname":"shetty", "gender":"Female"}]

#static method is used so we dont have to create a object of this class to call this method
    #we can directly call by classname.method name, also self parameter is not required for static method
    #Note- we can call variables directly by classname.variablename(mb)
    @staticmethod
    def getTestData(test_case_name):
            Dict = {}
            book = openpyxl.load_workbook("https:\\people.ey.com\\personal\\jeet_chakraborty1_gds_ey_com\\Documents\\Desktop\\TestData.xlsx?web=1")
            sheet = book.active
            for i in range(1, sheet.max_row + 1):  # to get rows
                if sheet.cell(row=i, column=1).value == test_case_name:

                    for j in range(2, sheet.max_column + 1):  # to get columns
                        # Dict["lastname"]="shetty
                        Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return[Dict]#returning the dictionary as list , as the fixture needs in list format